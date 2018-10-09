#!/usr/bin/env python

'''
handler for microsoft office documents
'''

import docx
from openpyxl import load_workbook
import datetime
import util as ut

def get_text_lines_xlsx(file):
    '''
    opens xslx doc
    extracts to array of lines
    needs work don't use
    '''
    wb = load_workbook(filename = file)
    ws = wb.active
    fulltext = []
    workingDuration = datetime.timedelta(0)
    incrementDuration = datetime.timedelta(minutes=1)
    workingParagraph = ''
    for row in range(2, ws.max_row + 1):
        rowData = {}
        rowData['tcIn'] = ws['B' + str(row)].value
        rowData['duration'] = ws['C' + str(row)].value
        rowData['txtOrig'] = ws['D' + str(row)].value
        rowData['txtTrans'] = ws['E' + str(row)].value
        timedText = {}
        if rowData['duration']:
            rowDuration = datetime.timedelta(hours=int(rowData['duration'][:2]), minutes=int(rowData['duration'][3:5]), seconds=int(rowData['duration'][6:8]))
            workingDuration = workingDuration + rowDuration
            workingParagraph = workingParagraph + ' ' + rowData['txtOrig']
            if workingDuration < incrementDuration:
                workingDuration = workingDuration + rowDuration
            else:
                minutes, remainder = divmod(workingDuration.seconds, 60)
                remainderPercentage = remainder / 60
                print(workingParagraph)
                print(workingDuration)
                allTheWordsInWorkingParagraph = workingParagraph.split()
                totalWordsInWorkingParagraph = len(allTheWordsInWorkingParagraph)
                wordsPerMinute = int(totalWordsInWorkingParagraph / (minutes + remainderPercentage))
                wordIndexIn = 0
                count = 1
                chunk = ''
                while wordIndexIn < totalWordsInWorkingParagraph:
                    wordIndexOut = wordIndexIn + wordsPerMinute
                    chunk = ' '.join(allTheWordsInWorkingParagraph[wordIndexIn:wordIndexOut])
                    timedText[count] = chunk
                    wordIndexIn = wordIndexOut
                    count = count+1
                print(timedText)
                workingDuration = datetime.timedelta(seconds=remainder)
                print(workingDuration)
    return fulltext

def get_text_lines_docx(filepath, kwargs):
    '''
    opens Word doc
    extracts text to array of lines
    '''
    if filepath.endswith('.docx'):
        doc = docx.Document(filepath)
        fulltext = []
        for para in doc.paragraphs:
            lines = []
            lines = make_lines(para.text, kwargs)
            for line in lines:
                fulltext.append(line)
    return fulltext

def make_lines(paragraph, kwargs):
    '''
    takes a word paragraph object and returns an array of lines
    '''
    charsInLine = kwargs.charsInLine
    paragraphLines = ['']
    oneLine = ''
    lines = paragraph.split('\n')
    for ln in lines: #need this to deal with <w:br/> line break tags, which aren't always mapped by openpyxl
        if not ln: #if line is blank/ carriage return
            paragraphLines.append(ln)
            oneLine = ''
        else:
            for word in ln.split():
                if not oneLine or len(oneLine) <= 1:
                    oneLine = word
                elif len(oneLine + ' ' + word) <= charsInLine:
                    oneLine = oneLine + ' ' + word
                else:
                    paragraphLines.append(oneLine)
                    oneLine = word
            if not oneLine in paragraphLines and not ln.replace("\n","") in paragraphLines:
                paragraphLines.append(oneLine)
    #print(paragraphLines)
    #print(len(paragraphLines))
    return paragraphLines

def main():
    '''
    do the thing
    '''
    kwargs = ut.dotdict({})
    kwargs.charsInLine = 90
    filepath = '/Volumes/2017_0215/LourdesPortillo/Transcript_LourdesPortillo_Eng_Public_CD_081618.docx'
    fulltext = get_text_lines_docx(filepath, kwargs)
    print(fulltext)

if __name__ == "__main__":
    main()
